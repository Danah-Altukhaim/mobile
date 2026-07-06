#!/usr/bin/env python3
"""Build faculty + equivalency reference data for packages/shared.

Run after build-catalog.py (reads its .cache/courses.json).
"""
import os, re, json
import openpyxl

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
DOCS = os.path.join(REPO, 'CCK Student Hub Docs')
SHARED = os.path.join(REPO, 'packages', 'shared', 'src')
CACHE = os.path.join(os.path.dirname(__file__), '.cache')

def ts_str(v):
    return 'null' if v is None else json.dumps(v, ensure_ascii=False)

def ts_arr(v):
    return '[' + ', '.join(ts_str(x) for x in v) + ']'

# ===========================================================================
# Faculty roster
# ===========================================================================
def norm_dept(raw):
    s = (raw or '').lower()
    if 'business' in s:
        return 'business'
    if 'foundation' in s or 'general education' in s or 'english' in s:
        return 'foundation'
    if 'advanced' in s or 'technology' in s or 'applied scien' in s:
        return 'advanced_technology'
    return 'business'

def norm_qual(raw):
    s = (raw or '').lower()
    if 'phd' in s:
        return 'PhD'
    if 'master' in s:
        return 'Master'
    if 'bachelor' in s:
        return 'Bachelor'
    return 'Master'

wb = openpyxl.load_workbook(os.path.join(DOCS, 'Acadamic Staff Data.xlsx'), data_only=True)
ws = wb.active
faculty = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name, title, qual, fac, ftpt, load, note = (list(row) + [None] * 7)[:7]
    if not name or not str(name).strip():
        continue
    name = re.sub(r'\s+', ' ', str(name).strip()).rstrip(',')
    title = (str(title).strip() if title else '')
    is_pt = 'p/t' in (str(ftpt) or '').lower() or 'part' in (str(title) or '').lower()
    load_n = None
    if isinstance(load, (int, float)):
        load_n = int(load)
    note_s = (str(note).strip() if note else '')
    is_hod = 'head of department' in note_s.lower()
    faculty.append({
        'name_en': name,
        'job_title': title,
        'qualification': norm_qual(qual),
        'department': norm_dept(fac),
        'employment': 'part_time' if is_pt else 'full_time',
        'is_hod': is_hod,
        'teaching_load': load_n,
    })

# ===========================================================================
# Instructor availability rules
# ===========================================================================
wb = openpyxl.load_workbook(os.path.join(DOCS, 'Per-instructor Availability.xlsx'), data_only=True)
ws = wb.active
avail = []
avail_notes = []
for row in ws.iter_rows(min_row=3, values_only=True):
    rank, band, days, hrs = (list(row) + [None] * 4)[:4]
    if not rank:
        continue
    if isinstance(days, (int, float)):
        avail.append({
            'rank': str(rank).strip(),
            'qualification_band': str(band).strip() if band else '',
            'days_available': int(days),
            'teaching_hours_per_week': int(hrs) if isinstance(hrs, (int, float)) else 0,
        })
    else:
        # footnote rule rows (e.g. "HOD max 12 hours") carried in column A only
        note = re.sub(r'\s+', ' ', str(rank).strip())
        if note and note not in avail_notes:
            avail_notes.append(note)

# ===========================================================================
# Faculty -> course offerings (match course names to catalog codes)
# ===========================================================================
with open(os.path.join(CACHE, 'courses.json')) as f:
    catalog = json.load(f)

def norm_name(s):
    s = (s or '').lower()
    s = s.replace('&', ' and ').replace('–', '-').replace('—', '-')
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    s = re.sub(r'\b(i)\b', '1', s)
    s = re.sub(r'\b(ii)\b', '2', s)
    s = re.sub(r'\b(iii)\b', '3', s)
    return re.sub(r'\s+', ' ', s)

# Course names in "List With Offered Courses By Staff" that use a shorthand
# title differing from the catalog's name_en.
COURSE_NAME_ALIASES = {
    'Cross-Platform Design': 'CST8117',
    'Database Systems and Concepts': 'CST8260',
    'Intro to Comp Programming using Python': 'CST8279',
}

name_index = {}
for c in catalog:
    name_index.setdefault(norm_name(c['name_en']), c['code'])

# --- instructor-name reconciliation across the two staff sheets ----------
# The offerings sheet writes names differently from the staff roster
# (titles Mr/Mrs/Dr, dots, and a few spelling variants). Normalising titles
# and punctuation joins most; the rest go through an explicit alias map.
# A leading title is "Dr"/"Mr"/... at a word boundary, then an optional dot
# and optional spaces — covers "Dr. Name", "Dr Name" and "Dr.Name".
TITLE_RE = re.compile(r'^(mr|mrs|ms|miss|dr|prof|professor)\b\.?\s*', re.IGNORECASE)

def norm_person(name):
    s = re.sub(r'\s+', ' ', str(name or '').strip()).rstrip(',')
    prev = None
    while prev != s:                       # strip stacked / repeated titles
        prev = s
        s = TITLE_RE.sub('', s).strip()
    # punctuation (hyphens, apostrophes) -> space so "Al-Qassar" == "Al Qassar"
    s = re.sub(r'[^a-z0-9]+', ' ', s.lower())
    return re.sub(r'\s+', ' ', s).strip()

faculty_name_index = {norm_person(m['name_en']): m['name_en'] for m in faculty}

# Offerings-sheet name (normalised) -> roster name, for genuine spelling
# variants that normalisation alone cannot reconcile (e.g. Ismaeil/Esmaiel).
FACULTY_NAME_ALIASES = {
    'ahmad esmaiel khaleel al qa ad': "Dr Ahmad Ismaeil Khaleel Al-Qa’ed",
}

def resolve_faculty(staff_raw):
    key = norm_person(staff_raw)
    if key in faculty_name_index:
        return faculty_name_index[key]
    if key in FACULTY_NAME_ALIASES:
        return FACULTY_NAME_ALIASES[key]
    # last resort: match on a shared surname + first-name token set
    toks = set(key.split())
    best = None
    for fkey, fname in faculty_name_index.items():
        ftoks = set(fkey.split())
        if len(toks & ftoks) >= max(2, min(len(toks), len(ftoks)) - 1):
            best = fname
            break
    return best

wb = openpyxl.load_workbook(os.path.join(DOCS, 'List With Offered Courses By Staff.xlsx'), data_only=True)
ws = wb.active
offerings = []
unmatched = set()
unmatched_staff = set()
seen = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    staff, course = (list(row) + [None, None])[:2]
    if not staff or not course:
        continue
    staff = re.sub(r'\s+', ' ', str(staff).strip()).rstrip(',')
    # normalize "Mrs. " / "Dr. " etc are kept; strip trailing punctuation
    course = re.sub(r'\s+', ' ', str(course).strip())
    code = COURSE_NAME_ALIASES.get(course) or name_index.get(norm_name(course))
    key = (staff, course)
    if key in seen:
        continue
    seen.add(key)
    if not code:
        unmatched.add(course)
    faculty_name = resolve_faculty(staff)
    if not faculty_name:
        unmatched_staff.add(staff)
    offerings.append({
        'instructor_name_en': staff,
        'faculty_name_en': faculty_name,
        'course_name_en': course,
        'course_code': code,
    })

# ===========================================================================
# CCK equivalency sheet
# ===========================================================================
CODE_RE = re.compile(r'\b([A-Z]{2,4})\s?(\d{4})[A-Z]?\b')
def norm_eq_code(raw):
    if not raw:
        return None
    m = CODE_RE.search(str(raw).upper().replace(' ', ''))
    return (m.group(1) + m.group(2)) if m else None

# Map columns by header label rather than fixed position — the source sheets
# have lost/gained columns between revisions (e.g. the 2026 "- Update" sheet
# dropped CCK "Equal to PAAET" and PAAET "Remarks"/"Rules"). A header lookup
# keeps the build correct instead of silently shifting every field one over.
def header_index(ws):
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {
        re.sub(r'\s+', ' ', str(h).strip()).lower(): i
        for i, h in enumerate(headers) if h is not None
    }

def cell(cells, idx, key):
    i = idx.get(key)
    return cells[i] if i is not None and i < len(cells) else None

def clean_str(v):
    return re.sub(r'\s+', ' ', str(v).strip()) if v and str(v).strip() else ''

wb = openpyxl.load_workbook(os.path.join(DOCS, 'Equivalency - Courses List - Update 2.xlsx'), data_only=True)
ws = wb['CCK']
cck_idx = header_index(ws)
equivalency = []
for row in ws.iter_rows(min_row=2, values_only=True):
    cells = list(row)
    cname = cell(cells, cck_idx, 'course name')
    ccode = cell(cells, cck_idx, 'course code')
    credit = cell(cells, cck_idx, 'credit')
    major = cell(cells, cck_idx, 'major')
    # "Equal to PAAET" was dropped in the 2026 update; keep the field, empty if absent.
    paaet = cell(cells, cck_idx, 'equal to paaet')
    remarks = cell(cells, cck_idx, 'remarks')
    if not cname and not ccode:
        continue
    equivalency.append({
        'cck_course_name': clean_str(cname),
        'cck_code': norm_eq_code(ccode),
        'cck_code_raw': clean_str(ccode) or None,
        'cck_credit': int(credit) if isinstance(credit, (int, float)) else None,
        'cck_major': clean_str(major),
        'paaet_program': clean_str(paaet),
        'remarks': clean_str(remarks) or None,
    })

# transferable-credit totals + rules pulled from the PAAET sheet's Rules column
# (absent in the 2026 update — `rules` stays empty when the column is gone).
ws2 = wb['PAAET']
paaet_idx = header_index(ws2)
rules = []
if 'rules' in paaet_idx:
    for row in ws2.iter_rows(min_row=2, values_only=True):
        rule = cell(list(row), paaet_idx, 'rules')
        if rule and str(rule).strip():
            r = re.sub(r'\s+', ' ', str(rule).strip())
            if r not in rules:
                rules.append(r)

# ===========================================================================
# PAAET equivalency sheet — reverse direction (one PAAET course -> CCK major)
# ===========================================================================
def fmt_credit(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)) and float(v) == int(v):
        return str(int(v))
    return re.sub(r'\s+', ' ', str(v).strip())

paaet_equiv = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    cells = list(row)
    cname = cell(cells, paaet_idx, 'course name')
    ccode = cell(cells, paaet_idx, 'course code')
    credit = cell(cells, paaet_idx, 'credit')
    program = cell(cells, paaet_idx, 'paaet course')
    major = cell(cells, paaet_idx, 'cck major')
    remarks = cell(cells, paaet_idx, 'remarks')
    if not cname and not ccode:
        continue
    paaet_equiv.append({
        'paaet_course_name': clean_str(cname),
        'paaet_code': clean_str(ccode),
        'credit': fmt_credit(credit),
        'paaet_program': clean_str(program),
        'cck_major': clean_str(major),
        'remarks': clean_str(remarks) or None,
    })

# ===========================================================================
# Emit faculty.ts
# ===========================================================================
lines = [
    '// AUTO-GENERATED from CCK Student Hub Docs — do not edit by hand.',
    '// Sources: Acadamic Staff Data.xlsx, Per-instructor Availability.xlsx,',
    '//          List With Offered Courses By Staff.xlsx',
    '// Regenerate: python3 packages/shared/scripts/build-reference.py',
    "import type {",
    '  FacultyMember,',
    '  InstructorAvailabilityRule,',
    '  FacultyCourseOffering,',
    "} from '../types/reference';",
    '',
    f'/** CCK academic staff roster — {len(faculty)} members. */',
    'export const FACULTY: FacultyMember[] = [',
]
for m in faculty:
    lines.append(
        '  { ' + ', '.join([
            f'name_en: {ts_str(m["name_en"])}',
            f'job_title: {ts_str(m["job_title"])}',
            f'qualification: {ts_str(m["qualification"])}',
            f'department: {ts_str(m["department"])}',
            f'employment: {ts_str(m["employment"])}',
            f'is_hod: {"true" if m["is_hod"] else "false"}',
            f'teaching_load: {m["teaching_load"] if m["teaching_load"] is not None else "null"}',
        ]) + ' },')
lines += ['];', '']
lines.append('/** Weekly availability / teaching-load rule by academic rank. */')
lines.append('export const INSTRUCTOR_AVAILABILITY: InstructorAvailabilityRule[] = [')
for a in avail:
    lines.append(
        '  { ' + ', '.join([
            f'rank: {ts_str(a["rank"])}',
            f'qualification_band: {ts_str(a["qualification_band"])}',
            f'days_available: {a["days_available"]}',
            f'teaching_hours_per_week: {a["teaching_hours_per_week"]}',
        ]) + ' },')
lines += ['];', '']
lines.append('/** Footnote rules from the Per-instructor Availability sheet. */')
lines.append('export const INSTRUCTOR_AVAILABILITY_NOTES: string[] = [')
for n in avail_notes:
    lines.append(f'  {ts_str(n)},')
lines += ['];', '']
lines.append(f'/** Instructor → course offerings ({len(offerings)} rows; '
             f'{sum(1 for o in offerings if o["course_code"])} matched to catalog codes). */')
lines.append('export const FACULTY_COURSE_OFFERINGS: FacultyCourseOffering[] = [')
for o in offerings:
    lines.append(
        '  { ' + ', '.join([
            f'instructor_name_en: {ts_str(o["instructor_name_en"])}',
            f'faculty_name_en: {ts_str(o["faculty_name_en"])}',
            f'course_name_en: {ts_str(o["course_name_en"])}',
            f'course_code: {ts_str(o["course_code"])}',
        ]) + ' },')
lines += [
    '];',
    '',
    '/** Course codes a given instructor (FacultyMember.name_en) is recorded',
    ' *  as offering — joined via the reconciled faculty_name_en field. */',
    'export function coursesForInstructor(name_en: string): string[] {',
    '  return FACULTY_COURSE_OFFERINGS.filter(',
    '    (o) => o.faculty_name_en === name_en && o.course_code,',
    '  ).map((o) => o.course_code as string);',
    '}',
    '',
]
with open(os.path.join(SHARED, 'data', 'faculty.ts'), 'w') as f:
    f.write('\n'.join(lines))

# ===========================================================================
# Emit equivalency.ts
# ===========================================================================
lines = [
    '// AUTO-GENERATED from CCK Student Hub Docs — do not edit by hand.',
    '// Source: Equivalency - Courses List - Update 2.xlsx ("CCK" + "PAAET" sheets)',
    '// Regenerate: python3 packages/shared/scripts/build-reference.py',
    "import type { EquivalencyEntry, PaaetEquivalencyEntry } from '../types/reference';",
    '',
    f'/** PAAET → CCK transfer-credit equivalency table ({len(equivalency)} entries). */',
    'export const EQUIVALENCY: EquivalencyEntry[] = [',
]
for e in equivalency:
    lines.append(
        '  { ' + ', '.join([
            f'cck_course_name: {ts_str(e["cck_course_name"])}',
            f'cck_code: {ts_str(e["cck_code"])}',
            f'cck_code_raw: {ts_str(e["cck_code_raw"])}',
            f'cck_credit: {e["cck_credit"] if e["cck_credit"] is not None else "null"}',
            f'cck_major: {ts_str(e["cck_major"])}',
            f'paaet_program: {ts_str(e["paaet_program"])}',
            f'remarks: {ts_str(e["remarks"])}',
        ]) + ' },')
lines += ['];', '']
lines.append(f'/** Per-course PAAET → CCK mappings — the reverse direction ({len(paaet_equiv)} entries). */')
lines.append('export const PAAET_EQUIVALENCY: PaaetEquivalencyEntry[] = [')
for e in paaet_equiv:
    lines.append(
        '  { ' + ', '.join([
            f'paaet_course_name: {ts_str(e["paaet_course_name"])}',
            f'paaet_code: {ts_str(e["paaet_code"])}',
            f'credit: {ts_str(e["credit"])}',
            f'paaet_program: {ts_str(e["paaet_program"])}',
            f'cck_major: {ts_str(e["cck_major"])}',
            f'remarks: {ts_str(e["remarks"])}',
        ]) + ' },')
lines += ['];', '']
lines.append('/** Transferable-credit totals and rules noted in the source sheet. */')
lines.append('export const EQUIVALENCY_RULES: string[] = [')
for r in rules:
    lines.append(f'  {ts_str(r)},')
lines += [
    '];',
    '',
    '/** Equivalency entries available to a transfer student from a PAAET program. */',
    'export function equivalencyForProgram(paaet_program: string): EquivalencyEntry[] {',
    '  return EQUIVALENCY.filter((e) =>',
    '    e.paaet_program.toLowerCase().includes(paaet_program.toLowerCase()),',
    '  );',
    '}',
    '',
    '/** Per-course PAAET → CCK mappings for a given PAAET diploma. */',
    'export function paaetEquivalencyForProgram(paaet_program: string): PaaetEquivalencyEntry[] {',
    '  return PAAET_EQUIVALENCY.filter((e) =>',
    '    e.paaet_program.toLowerCase().includes(paaet_program.toLowerCase()),',
    '  );',
    '}',
    '',
]
with open(os.path.join(SHARED, 'data', 'equivalency.ts'), 'w') as f:
    f.write('\n'.join(lines))

print(f"Faculty: {len(faculty)} ({sum(1 for m in faculty if m['is_hod'])} HODs, "
      f"{sum(1 for m in faculty if m['employment']=='full_time')} FT)")
print(f"Availability rules: {len(avail)}")
print(f"Course offerings: {len(offerings)} "
      f"({sum(1 for o in offerings if o['course_code'])} matched, {len(unmatched)} unmatched names)")
print(f"  unmatched courses: {sorted(unmatched)}")
print(f"Instructor name match: "
      f"{sum(1 for o in offerings if o['faculty_name_en'])}/{len(offerings)} resolved to faculty")
if unmatched_staff:
    print(f"  unmatched staff: {sorted(unmatched_staff)}")
print(f"Equivalency entries: {len(equivalency)}  | rules: {len(rules)}")
print("  -> wrote data/faculty.ts, data/equivalency.ts")
